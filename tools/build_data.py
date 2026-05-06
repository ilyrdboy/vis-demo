from __future__ import annotations

import json
import re
from collections import Counter
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[2]
APP = Path(__file__).resolve().parents[1]
OUT = APP / "src" / "data" / "processedData.js"
RECORDS_DIR = APP / "public" / "data" / "records"

PERIODS = [
    {"name": "先秦", "start": -900, "end": -221, "keys": ["先秦", "西周", "东周", "春秋", "战国", "燕国", "远古", "史前"]},
    {"name": "秦汉", "start": -221, "end": 220, "keys": ["秦", "汉", "西汉", "东汉", "两汉", "秦汉"]},
    {"name": "魏晋南北朝", "start": 220, "end": 581, "keys": ["魏晋", "北魏", "南北朝", "十六国", "晋"]},
    {"name": "隋唐五代", "start": 581, "end": 938, "keys": ["隋", "唐", "五代", "隋唐"]},
    {"name": "辽", "start": 938, "end": 1122, "keys": ["辽", "宋辽", "宋、辽"]},
    {"name": "金", "start": 1122, "end": 1215, "keys": ["金", "辽金"]},
    {"name": "元", "start": 1215, "end": 1368, "keys": ["元", "蒙古"]},
    {"name": "明", "start": 1368, "end": 1644, "keys": ["明"]},
    {"name": "清", "start": 1644, "end": 1912, "keys": ["清", "清末"]},
    {"name": "民国", "start": 1912, "end": 1949, "keys": ["民国", "近代"]},
]

DATASETS = [
    {"key": "water", "filename": "02水系 - 总数据和各朝代数据.xlsx", "label": "水系", "titleCols": ["河流/水系名称", "水系总体描述的原文", "图名"], "rawCols": ["水系总体描述的原文"], "sourceCols": ["出处"]},
    {"key": "climate", "filename": "03气候 - 总数据和各朝代数据.xlsx", "label": "气候", "titleCols": ["气候总体描述原文", "温度/冷暖"], "rawCols": ["气候总体描述原文"], "sourceCols": ["出处"]},
    {"key": "vegetation", "filename": "04植被 - 总数据和各朝代数据.xlsx", "label": "植被", "titleCols": ["植被类型", "植被分布", "植被总体描述原文"], "rawCols": ["植被总体描述原文"], "sourceCols": ["出处"]},
    {"key": "disaster", "filename": "05灾害 - 总数据和各朝代数据.xlsx", "label": "灾害", "titleCols": ["灾害类型", "灾害时间", "灾害总体描述原文"], "rawCols": ["灾害总体描述原文"], "sourceCols": ["出处"]},
    {"key": "governance", "filename": "07建制沿革 - 总数据和各朝代数据.xlsx", "label": "建制沿革", "titleCols": ["行政划分", "机构设置", "建制沿革总体描述原文"], "rawCols": ["建制沿革总体描述原文"], "sourceCols": ["出处"]},
    {"key": "keyBuildings", "filename": "09重点建筑 - 总数据和各朝代数据.xlsx", "label": "重点建筑", "titleCols": ["建筑名称", "建筑类别", "图名"], "rawCols": ["建筑描述原文"], "sourceCols": ["出处"]},
    {"key": "buildings", "filename": "10其他建筑 - 总数据和各朝代数据.xlsx", "label": "其他建筑", "titleCols": ["建筑名称", "建筑类别", "图名"], "rawCols": ["城市建筑总体描述"], "sourceCols": ["出处"]},
    {"key": "population", "filename": "11人口 - 总数据和各朝代数据.xlsx", "label": "人口", "titleCols": ["人口数量", "人口总体描述原文", "原文"], "rawCols": ["人口总体描述原文", "原文"], "sourceCols": ["出处"]},
    {"key": "culture", "filename": "13文化 - 总数据和各朝代数据.xlsx", "label": "文化", "titleCols": ["文化", "文化总体描述原文", "原文"], "rawCols": ["文化总体描述原文", "原文", "文化"], "sourceCols": ["出处"]},
    {"key": "commerce", "filename": "14商业手工业 -总数据和各朝代数据.xlsx", "label": "商业手工业", "titleCols": ["商业手工业类型", "商业手工业总体描述原文"], "rawCols": ["商业手工业总体描述原文", "原文"], "sourceCols": ["出处"]},
    {"key": "products", "filename": "15物产 - 总数据和各朝代数据.xlsx", "label": "物产", "titleCols": ["物产类型", "物产", "物产总体描述原文"], "rawCols": ["物产总体描述原文", "原文"], "sourceCols": ["出处"]},
    {"key": "traffic", "filename": "16交通 - 总数据和各朝代数据.xlsx", "label": "交通", "titleCols": ["交通方式", "交通工具", "交通要道", "交通总体描述原文"], "rawCols": ["交通总体描述原文", "原文"], "sourceCols": ["出处"]},
    {"key": "events", "filename": "17事件 - 总数据和各朝代数据.xlsx", "label": "事件", "titleCols": ["事件概括", "事件原文"], "rawCols": ["事件原文"], "sourceCols": ["出处"]},
    {"key": "wars", "filename": "18战争 - 总数据和各朝代数据.xlsx", "label": "战争", "titleCols": ["战争名称", "战争描述原文"], "rawCols": ["战争描述原文"], "sourceCols": ["出处"]},
    {"key": "people", "filename": "19人物 - 总数据和各朝代数据.xlsx", "label": "人物", "titleCols": ["姓名", "身份"], "rawCols": ["该人物在北京的活动", "人物介绍"], "sourceCols": ["出处"]},
]

TURNING_POINT_RULES = {
    "yan-cloud": {"period": "辽", "datasets": ["events", "governance", "wars"], "terms": ["燕云", "南京", "辽", "十六州"]},
    "jin-capital": {"period": "金", "datasets": ["keyBuildings", "water", "population"], "terms": ["海陵", "迁都", "中都", "金口河", "卢沟"]},
    "yuan-dadu": {"period": "元", "datasets": ["wars", "events", "traffic"], "terms": ["蒙古", "中都", "大都", "通惠河", "漕运"]},
    "ming-beiping": {"period": "明", "datasets": ["wars", "governance", "population"], "terms": ["徐达", "大都", "北平", "迁徙", "洪武"]},
    "modern-capital": {"period": "民国", "datasets": ["events", "population", "traffic", "culture", "people"], "terms": ["北京", "首都", "中央", "人民政府", "文化"]},
}

SERIES_DATASET_KEYS = ["disaster", "wars", "population", "traffic"]
EVENT_DATASET_KEYS = ["events", "wars", "disaster", "population", "traffic", "governance", "water", "people"]

KNOWN_PLACES = [
    {
        "id": "beijing",
        "name": "北京",
        "modernName": "北京市",
        "aliases": ["北京", "燕京", "蓟城", "幽州", "南京", "中都", "大都", "北平", "京师", "顺天府"],
        "kind": "都城核心",
        "x": 52,
        "y": 50,
        "note": "北京历史都城空间的核心节点，承载政权、人口和交通功能转换。",
    },
    {
        "id": "datong",
        "name": "大同",
        "modernName": "山西大同",
        "aliases": ["大同", "云中", "燕云", "代北"],
        "kind": "北方门户",
        "x": 28,
        "y": 34,
        "note": "连接燕云、山西与草原交通的西北门户。",
    },
    {
        "id": "shangdu",
        "name": "上都",
        "modernName": "内蒙古锡林郭勒",
        "aliases": ["上都", "开平", "蒙古", "草原"],
        "kind": "草原腹地",
        "x": 42,
        "y": 18,
        "note": "元代两都与草原政治网络的重要方向。",
    },
    {
        "id": "liaodong",
        "name": "辽东",
        "modernName": "辽宁一带",
        "aliases": ["辽东", "女真", "契丹", "东北"],
        "kind": "东北通道",
        "x": 80,
        "y": 30,
        "note": "契丹、女真与东北交通军事网络的指向。",
    },
    {
        "id": "tongzhou",
        "name": "通州",
        "modernName": "北京通州",
        "aliases": ["通州", "张家湾", "潞河"],
        "kind": "漕运门户",
        "x": 63,
        "y": 55,
        "note": "运河入京与物资转运的重要节点。",
    },
    {
        "id": "tianjin",
        "name": "天津",
        "modernName": "天津",
        "aliases": ["天津", "直沽", "海河"],
        "kind": "海河口岸",
        "x": 70,
        "y": 73,
        "note": "连接京师、海河与海运的近现代门户。",
    },
    {
        "id": "kaifeng",
        "name": "开封",
        "modernName": "河南开封",
        "aliases": ["开封", "汴梁", "河南", "中原"],
        "kind": "中原方向",
        "x": 36,
        "y": 82,
        "note": "从中原王朝、南北交通看北京地位变化的参照方向。",
    },
    {
        "id": "jiangnan",
        "name": "江南",
        "modernName": "江南地区",
        "aliases": ["江南", "杭州", "苏州", "应天", "南方"],
        "kind": "粮运腹地",
        "x": 55,
        "y": 93,
        "note": "漕运、财富和文化输入的重要区域。",
    },
]

KNOWN_WATERWAYS = [
    {"id": "yongding", "name": "永定河", "aliases": ["永定河", "桑乾河", "卢沟河", "无定河"], "source": "西山水系", "function": "防洪与城西通道"},
    {"id": "tonghui", "name": "通惠河", "aliases": ["通惠河", "坝河", "惠河"], "source": "元明漕运", "function": "大都漕运入城"},
    {"id": "canal", "name": "大运河", "aliases": ["大运河", "运河", "漕河", "京杭运河"], "source": "南北水运", "function": "国家尺度物资运输"},
    {"id": "gaoliang", "name": "高梁河", "aliases": ["高梁河", "高粱河"], "source": "西北水系", "function": "城西北水源与交通"},
    {"id": "jinkou", "name": "金口河", "aliases": ["金口河", "金口"], "source": "金元水利", "function": "引水与漕运尝试"},
    {"id": "chaobai", "name": "潮白河", "aliases": ["潮白河", "白河", "潮河"], "source": "东北水系", "function": "京东水文联系"},
]

ROUTE_RULES = [
    {"source": "beijing", "target": "datong", "label": "燕云-山西通道", "aliases": ["燕云", "大同", "云中", "代北", "山西"]},
    {"source": "beijing", "target": "shangdu", "label": "两都草原通道", "aliases": ["上都", "开平", "蒙古", "草原", "元上都"]},
    {"source": "beijing", "target": "liaodong", "label": "东北军政通道", "aliases": ["辽东", "女真", "契丹", "东北", "辽"]},
    {"source": "beijing", "target": "tongzhou", "label": "通州漕运门户", "aliases": ["通州", "张家湾", "潞河", "漕运", "运河"]},
    {"source": "tongzhou", "target": "tianjin", "label": "海河转运通道", "aliases": ["天津", "直沽", "海河", "海运"]},
    {"source": "beijing", "target": "kaifeng", "label": "中原政治通道", "aliases": ["开封", "汴梁", "中原", "河南"]},
    {"source": "tongzhou", "target": "jiangnan", "label": "江南粮运通道", "aliases": ["江南", "杭州", "苏州", "南方", "京杭运河", "大运河"]},
]

PEOPLE_ALLOWLIST = {
    "徐达", "海陵王", "完颜亮", "成吉思汗", "忽必烈", "朱元璋", "朱棣", "周武王", "周厉王",
    "康熙", "乾隆", "光绪", "慈禧", "袁世凯", "孙中山", "毛泽东", "刘秉忠", "郭守敬",
}

STATUS_TERMS = {
    "开凿": ["开凿", "疏浚", "修浚", "引水"],
    "改道": ["改道", "迁徙", "徙", "北移", "南移"],
    "淤积": ["淤积", "壅塞", "淤塞", "泥沙"],
    "治理": ["治理", "修筑", "堤", "闸", "减水"],
    "废弃": ["废弃", "废", "停用", "断流"],
    "漕运": ["漕运", "运粮", "转运", "粮运"],
    "灾害": ["水灾", "洪水", "泛滥", "决口"],
}

STOPWORDS = {
    "总体描述", "北京地区", "北京市", "时期", "出处", "数据", "相关", "其中", "进行", "已经", "城市",
    "但是", "及其", "由于", "因此", "所以", "同时", "此时", "此后", "形成", "出现", "开始", "具有",
    "记载", "说明", "发展", "影响", "变化", "重要", "主要", "相关", "可能", "成为", "属于",
    "正月", "二月", "三月", "四月", "五月", "六月", "七月", "八月", "九月", "十月", "十一月", "十二月",
    "北京市社会", "北京市社科", "社会科学", "学研究所", "北京历史", "历史纪年", "编写组", "出版社",
    "人民出版社", "北京出版社", "北京燕山", "研究成果", "本书", "资料", "原文",
}

HISTORICAL_ALLOWLIST = {
    "燕云十六州", "幽州", "蓟城", "燕京", "南京", "中都", "大都", "北平", "京师", "顺天府",
    "燕国", "辽", "金", "元", "明", "清", "民国", "蒙古", "契丹", "女真",
    "永定河", "卢沟河", "金口河", "通惠河", "高梁河", "漕运", "运河", "水灾", "旱灾",
    "徐达", "海陵王", "成吉思汗", "忽必烈", "朱元璋", "朱棣", "周武王", "周厉王",
    "迁都", "改制", "分封制", "郡县制", "节度使", "都城", "宫殿", "城垣", "战争", "灾害",
    "粮食", "石器", "蚌器", "家畜", "金属", "穷汉市", "畜牧", "陶器", "铁器", "铜器",
    "果品", "药材", "煤炭", "玉泉", "稻米", "麦", "谷", "粟", "马", "牛", "羊",
}

HISTORICAL_SUFFIXES = (
    "河", "渠", "城", "都", "京", "府", "州", "郡", "县", "门", "桥", "寺", "宫", "殿", "坛", "陵",
    "王", "帝", "宗", "汗", "军", "兵", "乱", "战", "役", "灾", "制", "法", "府", "司", "院", "营",
)

HISTORICAL_TERMS = (
    "迁都", "攻占", "围破", "改建", "设立", "开凿", "改道", "淤积", "漕运", "分封", "郡县",
    "节度使", "人口", "宫殿", "都城", "政权", "战争", "灾害", "水系", "交通", "建制", "燕云",
    "粮食", "物产", "家畜", "畜牧", "金属", "陶器", "石器",
)


def text(value: object) -> str:
    return str(value or "").strip()


def normalize_period(value: object) -> str:
    value_text = text(value)
    if not value_text or value_text in {"0", "时期", "公元纪年"}:
        return "未标注"
    for period in PERIODS:
        if any(key in value_text for key in period["keys"]):
            return period["name"]
    return "其他"


def merge_header_rows(parent: list[str], child: list[str]) -> list[str]:
    headers = []
    for index, value in enumerate(child):
        if value:
            headers.append(value)
            continue
        parent_value = parent[index] if index < len(parent) else ""
        headers.append(parent_value)
    return headers


def is_subheader_row(parent: list[str], child: list[str]) -> bool:
    if not parent or not child:
        return False
    if "时期" not in parent and parent[0] != "时期":
        return False
    child_terms = {"原文", "出处", "图名", "类型", "交通工具", "交通要道", "物产类型", "商业手工业类型"}
    child_hits = sum(1 for cell in child if cell in child_terms or cell.endswith("类型"))
    return child_hits >= 2


def find_header_row(sheet) -> tuple[int, list[str]]:
    preview_rows = [
        [text(cell) for cell in row]
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 8), values_only=True)
    ]
    for index in range(1, len(preview_rows)):
        parent = preview_rows[index - 1]
        child = preview_rows[index]
        if is_subheader_row(parent, child):
            headers = merge_header_rows(parent, child)
            if headers and not headers[0]:
                headers[0] = "时期"
            return index + 1, headers

    best_row = 1
    best_headers: list[str] = []
    best_score = -1
    for row_index, headers in enumerate(preview_rows, start=1):
        score = sum(1 for cell in headers if cell)
        score += sum(2 for cell in headers if cell in {"时期", "出处", "事件概括", "姓名", "建筑名称"})
        if score > best_score:
            best_row = row_index
            best_headers = headers
            best_score = score
    if "时期" not in best_headers and best_headers:
        best_headers[0] = "时期"
    return best_row, best_headers


def row_to_record(headers: list[str], cells: list[str]) -> dict[str, str]:
    record = {}
    for index, value in enumerate(cells):
        header = headers[index] if index < len(headers) and headers[index] else f"列{index + 1}"
        if header in record:
            header = f"{header}_{index + 1}"
        record[header] = value
    return record


def pick_first(record: dict[str, str], candidates: list[str], fallback_indexes: list[int] | None = None) -> str:
    for candidate in candidates:
        for key, value in record.items():
            if candidate == key or candidate in key:
                if value:
                    return value
    if fallback_indexes:
        values = list(record.values())
        for index in fallback_indexes:
            if index < len(values) and values[index]:
                return values[index]
    return ""


def pick_fallback_text(record: dict[str, str], fallback_indexes: list[int]) -> str:
    values = list(record.items())
    excluded_terms = ("出处", "来源", "图片", "图像", "水系图", "建筑图")
    for index in fallback_indexes:
        if index < len(values):
            key, value = values[index]
            if value and not any(term in key for term in excluded_terms):
                return value
    return ""


def is_historical_keyword(token: str) -> bool:
    if token in STOPWORDS:
        return False
    if len(token) < 2 or len(token) > 8:
        return False
    if token in HISTORICAL_ALLOWLIST:
        return True
    if any(term in token for term in HISTORICAL_TERMS):
        return True
    if token.endswith(HISTORICAL_SUFFIXES):
        return True
    if re.search(r"(辽|金|元|明|清|燕|蓟|宋|唐|隋|秦|汉|蒙古|契丹|女真)", token):
        return True
    return False


def keywords_from_text(raw: str, limit: int = 8) -> list[str]:
    counter = Counter()
    candidates = set(HISTORICAL_ALLOWLIST)
    candidates.update(re.findall(r"[\u4e00-\u9fa5]{2,8}", raw))
    for token in candidates:
        if token in raw and is_historical_keyword(token):
            counter[token] += 1
    return [token for token, _ in sorted(counter.items(), key=lambda item: (-item[1], len(item[0]), item[0]))[:limit]]


def sheet_profiles(workbook) -> list[dict]:
    return [
        {"name": sheet.title, "rows": sheet.max_row, "cols": sheet.max_column}
        for sheet in workbook.worksheets
        if sheet.max_row > 1
    ]


def score_sample(record: dict, rule: dict | None = None) -> int:
    haystack = f"{record.get('period', '')} {record.get('title', '')} {record.get('rawText', '')}"
    score = min(len(record.get("rawText", "")), 600) // 60
    if rule:
        if record.get("period") == rule["period"]:
            score += 8
            score += sum(4 for term in rule["terms"] if term and term in haystack)
    return score


def browse_score(record: dict) -> int:
    score = score_sample(record)
    if record.get("period") not in {"未标注", "其他"}:
        score += 6
    if record.get("source"):
        score += 3
    if record.get("title") and record.get("title") not in {"原文", "记录"}:
        score += 2
    return score


def year_from_text(raw: str) -> int | None:
    candidates: list[int] = []
    for match in re.finditer(r"(公元前|前)\s*(\d{1,4})\s*年", raw):
        candidates.append(-int(match.group(2)))
    for match in re.finditer(r"(?<!\d)(\d{3,4})\s*年", raw):
        year = int(match.group(1))
        if 900 <= year <= 2026:
            candidates.append(year)
    return candidates[0] if candidates else None


def period_midpoint(period_name: str) -> int | None:
    for period in PERIODS:
        if period["name"] == period_name:
            return round((period["start"] + period["end"]) / 2)
    return None


def best_year(record: dict) -> int | None:
    haystack = f"{record.get('title', '')} {record.get('rawText', '')}"
    return year_from_text(haystack) or period_midpoint(record.get("period", ""))


def matches_any(raw: str, aliases: list[str]) -> bool:
    return any(alias and alias in raw for alias in aliases)


def extract_places(raw: str) -> list[str]:
    places = [place["id"] for place in KNOWN_PLACES if matches_any(raw, place["aliases"])]
    return places or (["beijing"] if "北京" in raw else [])


def extract_people(raw: str) -> list[str]:
    found = {person for person in PEOPLE_ALLOWLIST if person in raw}
    found.update(re.findall(r"[\u4e00-\u9fa5]{2,4}(?:帝|王|宗|汗)", raw))
    return sorted(found)[:6]


def classify_event_type(record: dict) -> str:
    dataset = record.get("dataset", "")
    raw = f"{record.get('title', '')} {record.get('rawText', '')}"
    if dataset == "wars" or any(term in raw for term in ["战争", "攻占", "围", "破", "兵", "军"]):
        return "战争"
    if dataset == "disaster" or any(term in raw for term in ["灾", "旱", "水", "震", "疫", "蝗"]):
        return "灾害"
    if any(term in raw for term in ["迁都", "定都", "建都", "改为", "设置", "设立"]):
        return "建制"
    if dataset == "traffic" or any(term in raw for term in ["交通", "漕运", "道路", "铁路", "运河"]):
        return "交通"
    if dataset == "population" or any(term in raw for term in ["人口", "迁徙", "户", "居民"]):
        return "人口"
    if dataset == "people":
        return "人物"
    return "事件"


def collect_dataset(config: dict) -> dict:
    path = ROOT / config["filename"]
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    header_row, headers = find_header_row(sheet)
    period_counts = Counter()
    keyword_counts = Counter()
    samples = []
    trace_records = []
    empty_rows = 0
    records = 0

    for row_index, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        cells = [text(cell) for cell in row]
        if not any(cells):
            empty_rows += 1
            continue

        row_record = row_to_record(headers, cells)
        period = normalize_period(pick_first(row_record, ["时期", "公元纪年"], [0]))
        title = pick_first(row_record, config["titleCols"], [1, 2, 3])
        raw = pick_first(row_record, config["rawCols"]) or pick_fallback_text(row_record, [1, 2, 3, 4])
        source = pick_first(row_record, config["sourceCols"])
        combined = " ".join(value for value in [title, raw] if value)
        keywords = keywords_from_text(combined)

        period_counts[period] += 1
        records += 1
        for keyword in keywords:
            keyword_counts[keyword] += 1

        trace = {
            "dataset": config["key"],
            "datasetLabel": config["label"],
            "filename": config["filename"],
            "sheet": sheet.title,
            "row": row_index,
            "period": period,
            "title": title or f"{config['label']}记录 {row_index}",
            "rawText": raw if raw else combined,
            "source": source,
            "keywords": keywords,
        }
        trace_records.append(trace)
        if len(samples) < 6:
            samples.append(trace)

    quality = {
        "emptyRows": empty_rows,
        "missingPeriod": period_counts.get("未标注", 0),
        "missingPeriodRate": round(period_counts.get("未标注", 0) / max(records, 1), 3),
    }

    workbook.close()
    browse_records = sorted(trace_records, key=browse_score, reverse=True)[:180]
    return {
        "key": config["key"],
        "label": config["label"],
        "filename": config["filename"],
        "sizeMb": round(path.stat().st_size / 1024 / 1024, 2),
        "records": records,
        "rows": sheet.max_row,
        "cols": sheet.max_column,
        "sheets": sheet_profiles(openpyxl.load_workbook(path, read_only=True, data_only=True)),
        "periodCounts": dict(period_counts),
        "topKeywords": [{"name": key, "value": value} for key, value in keyword_counts.most_common(10)],
        "samples": samples,
        "browseRecords": browse_records,
        "quality": quality,
        "_traceRecords": trace_records,
    }


def build_period_matrix(datasets: list[dict]) -> list[dict]:
    order = [period["name"] for period in PERIODS] + ["其他", "未标注"]
    matrix = []
    for period in order:
        row = {"period": period}
        for dataset in datasets:
            row[dataset["key"]] = dataset["periodCounts"].get(period, 0)
        matrix.append(row)
    return matrix


def build_focus_series(datasets: list[dict]) -> list[dict]:
    focus_keys = ["water", "disaster", "governance", "population", "traffic", "events", "wars", "people"]
    return [
        {
            "key": dataset["key"],
            "label": dataset["label"],
            "total": dataset["records"],
            "periodCounts": dataset["periodCounts"],
            "topKeywords": dataset["topKeywords"][:6],
        }
        for dataset in datasets
        if dataset["key"] in focus_keys
    ]


def build_evidence_index(datasets: list[dict]) -> dict:
    by_key = {dataset["key"]: dataset for dataset in datasets}
    result = {}
    for point_id, rule in TURNING_POINT_RULES.items():
        records = []
        for dataset_key in rule["datasets"]:
            dataset = by_key.get(dataset_key)
            if not dataset:
                continue
            ranked = sorted(dataset["_traceRecords"], key=lambda record: score_sample(record, rule), reverse=True)
            records.extend(ranked[:3])
        result[point_id] = sorted(records, key=lambda record: score_sample(record, rule), reverse=True)[:8]
    return result


def build_event_records(datasets: list[dict]) -> list[dict]:
    records = []
    for dataset in datasets:
        if dataset["key"] not in EVENT_DATASET_KEYS:
            continue
        for record in dataset["_traceRecords"]:
            year = best_year(record)
            if year is None:
                continue
            haystack = f"{record.get('title', '')} {record.get('rawText', '')}"
            score = score_sample(record)
            score += 6 if -900 <= year <= 1949 else 0
            score += len(extract_places(haystack)) * 2
            score += len(extract_people(haystack))
            records.append(
                {
                    "id": f"{record['dataset']}-{record['row']}",
                    "dataset": record["dataset"],
                    "datasetLabel": record["datasetLabel"],
                    "filename": record["filename"],
                    "sheet": record["sheet"],
                    "row": record["row"],
                    "period": record["period"],
                    "year": year,
                    "type": classify_event_type(record),
                    "title": record["title"],
                    "rawText": record["rawText"],
                    "source": record["source"],
                    "keywords": record["keywords"],
                    "places": extract_places(haystack),
                    "people": extract_people(haystack),
                    "_score": score,
                }
            )
    records = sorted(records, key=lambda item: (item["_score"], item["year"]), reverse=True)
    public = []
    for item in records[:160]:
        clean = dict(item)
        clean.pop("_score", None)
        public.append(clean)
    return sorted(public, key=lambda item: item["year"])


def build_year_series(event_records: list[dict]) -> list[dict]:
    buckets: dict[int, Counter] = {}
    for record in event_records:
        year = record["year"]
        bucket = (year // 50) * 50 if year >= 0 else -((-year + 49) // 50) * 50
        buckets.setdefault(bucket, Counter())
        buckets[bucket][record["type"]] += 1
        buckets[bucket]["total"] += 1
    return [
        {
            "year": year,
            "label": f"前{abs(year)}" if year < 0 else str(year),
            "total": counts["total"],
            "wars": counts["战争"],
            "disasters": counts["灾害"],
            "governance": counts["建制"],
            "traffic": counts["交通"],
            "population": counts["人口"],
            "events": counts["事件"] + counts["人物"],
        }
        for year, counts in sorted(buckets.items())
    ]


def build_entity_index(event_records: list[dict]) -> dict:
    place_counts = Counter()
    person_counts = Counter()
    keyword_counts = Counter()
    for record in event_records:
        for place in record["places"]:
            place_counts[place] += 1
        for person in record["people"]:
            person_counts[person] += 1
        for keyword in record["keywords"]:
            keyword_counts[keyword] += 1

    places = []
    for place in KNOWN_PLACES:
        item = dict(place)
        item["value"] = place_counts.get(place["id"], 0)
        places.append(item)

    return {
        "places": places,
        "people": [{"name": key, "value": value} for key, value in person_counts.most_common(20)],
        "keywords": [{"name": key, "value": value} for key, value in keyword_counts.most_common(24)],
    }


def build_geo_network(event_records: list[dict]) -> dict:
    route_counts = Counter()
    route_samples: dict[str, dict] = {}
    for record in event_records:
        haystack = f"{record.get('title', '')} {record.get('rawText', '')}"
        for rule in ROUTE_RULES:
            if matches_any(haystack, rule["aliases"]):
                key = f"{rule['source']}->{rule['target']}"
                route_counts[key] += 1
                route_samples.setdefault(key, record)

    routes = []
    for rule in ROUTE_RULES:
        key = f"{rule['source']}->{rule['target']}"
        sample = route_samples.get(key)
        routes.append(
            {
                "source": rule["source"],
                "target": rule["target"],
                "label": rule["label"],
                "value": route_counts.get(key, 0),
                "sample": {
                    "title": sample["title"],
                    "period": sample["period"],
                    "year": sample["year"],
                    "datasetLabel": sample["datasetLabel"],
                    "rawText": sample["rawText"],
                }
                if sample
                else None,
            }
        )
    return {"places": KNOWN_PLACES, "routes": [route for route in routes if route["value"] > 0]}


def build_water_lifelines(datasets: list[dict]) -> list[dict]:
    water_dataset = next((dataset for dataset in datasets if dataset["key"] == "water"), None)
    if not water_dataset:
        return []

    result = []
    for waterway in KNOWN_WATERWAYS:
        records = []
        status_counts = Counter()
        period_counts = Counter()
        for record in water_dataset["_traceRecords"]:
            haystack = f"{record.get('title', '')} {record.get('rawText', '')}"
            if not matches_any(haystack, waterway["aliases"]):
                continue
            statuses = [
                status
                for status, terms in STATUS_TERMS.items()
                if any(term in haystack for term in terms)
            ] or ["记载"]
            for status in statuses:
                status_counts[status] += 1
            period_counts[record["period"]] += 1
            records.append(
                {
                    "year": best_year(record) or period_midpoint(record["period"]) or 0,
                    "period": record["period"],
                    "status": statuses[0],
                    "title": record["title"],
                    "rawText": record["rawText"],
                    "source": record["source"],
                    "row": record["row"],
                }
            )
        records = sorted(records, key=lambda item: item["year"])
        result.append(
            {
                "id": waterway["id"],
                "name": waterway["name"],
                "source": waterway["source"],
                "function": waterway["function"],
                "records": len(records),
                "statusCounts": dict(status_counts),
                "periodCounts": dict(period_counts),
                "events": records[:80],
            }
        )
    return [item for item in result if item["records"] > 0]


def main():
    RECORDS_DIR.mkdir(parents=True, exist_ok=True)
    datasets = [collect_dataset(config) for config in DATASETS]
    evidence_index = build_evidence_index(datasets)
    event_records = build_event_records(datasets)
    public_datasets = []
    for dataset in datasets:
        records_path = RECORDS_DIR / f"{dataset['key']}.json"
        records_path.write_text(
            json.dumps(dataset["_traceRecords"], ensure_ascii=False, separators=(",", ":")),
            encoding="utf-8",
        )
        clean = dict(dataset)
        clean.pop("_traceRecords", None)
        clean["recordsUrl"] = f"data/records/{dataset['key']}.json"
        public_datasets.append(clean)

    payload = {
        "generatedFrom": "D:/vis/*.xlsx",
        "datasets": public_datasets,
        "periodMatrix": build_period_matrix(public_datasets),
        "focusSeries": build_focus_series(public_datasets),
        "evidenceIndex": evidence_index,
        "eventRecords": event_records,
        "yearSeries": build_year_series(event_records),
        "entityIndex": build_entity_index(event_records),
        "geoNetwork": build_geo_network(event_records),
        "waterLifelines": build_water_lifelines(datasets),
        "totals": {
            "datasets": len(public_datasets),
            "records": sum(item["records"] for item in public_datasets),
            "sizeMb": round(sum(item["sizeMb"] for item in public_datasets), 2),
            "events": len(event_records),
        },
    }
    OUT.write_text(
        "export const processedData = "
        + json.dumps(payload, ensure_ascii=False, indent=2)
        + "\n",
        encoding="utf-8",
    )
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
